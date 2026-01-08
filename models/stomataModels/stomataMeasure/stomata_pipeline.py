# stomata_pipeline.py
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List, Union

import cv2
import numpy as np
import pandas as pd

from ultralytics import YOLO
from segment_anything import sam_model_registry, SamPredictor

from shapely.geometry import Polygon
from rasterio.features import rasterize
from affine import Affine

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ------------------------- Config dataclasses ------------------------- #

@dataclass(frozen=True)
class PipelineConfig:
    imgsz: int = 2560
    device: Union[int, str] = 0  # 0 or "cuda:0" or "cpu" or "mps"
    conf: float = 0.25
    iou: float = 0.7
    amp: bool = False
    rect: bool = True
    verbose: bool = False


@dataclass(frozen=True)
class ScaleConfig:
    um_per_px: float  # micrometers per pixel


@dataclass(frozen=True)
class OutputConfig:
    out_dir: Union[str, Path]
    save_overlay: bool = True
    save_excel: bool = True
    overlay_suffix: str = "_overlay.png"
    excel_suffix: str = "_results.xlsx"


@dataclass
class ImageResult:
    image_name: str
    summary: Dict[str, Any]
    instances_df: pd.DataFrame
    overlay_bgr: np.ndarray
    overlay_path: Optional[str] = None
    excel_path: Optional[str] = None
    elapsed_sec: Optional[float] = None


# ------------------------- Device helper ------------------------- #

def normalize_device(device: Union[int, str]) -> Tuple[Union[int, str], str]:
    """
    Returns: (yolo_device, sam_device_str)
    - Ultralytics YOLO accepts int (0) or "cuda:0"/"cpu"/"mps"
    - SAM .to(...) accepts strings like "cuda:0"/"cpu"/"mps"
    """
    if isinstance(device, int):
        return device, f"cuda:{device}"

    d = str(device).strip().lower()
    if d in {"cpu", "mps"}:
        return d, d
    if d.startswith("cuda"):
        return d, d if ":" in d else "cuda:0"
    if d.isdigit():
        di = int(d)
        return di, f"cuda:{di}"

    return device, str(device)


# ------------------------- Helpers (pure functions) ------------------------- #

def load_bgr(path: Union[str, Path]) -> np.ndarray:
    img = cv2.imread(str(path), cv2.IMREAD_UNCHANGED)
    if img is None:
        raise FileNotFoundError(f"Could not read image: {path}")

    if img.ndim == 2:
        img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    elif img.shape[2] == 4:
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)
    return img


def obb_polygon_mask(pts: np.ndarray, h: int, w: int) -> np.ndarray:
    poly = Polygon([(float(x), float(y)) for x, y in pts])
    if (not poly.is_valid) or poly.area <= 0:
        return np.zeros((h, w), dtype=bool)

    m = rasterize(
        [(poly, 1)],
        out_shape=(h, w),
        transform=Affine.identity(),
        fill=0,
        all_touched=False,
        dtype="uint8",
    )
    return m.astype(bool)


def mask_to_mbr(mask_bool: np.ndarray):
    ys, xs = np.where(mask_bool)
    if len(xs) < 10:
        return None, 0.0, 0.0, None

    pts = np.stack([xs, ys], axis=1).astype(np.float32)
    rect = cv2.minAreaRect(pts)
    (cx, cy), (rw, rh), _ = rect
    box = cv2.boxPoints(rect).astype(np.float32)

    length = float(max(rw, rh))
    width = float(min(rw, rh))
    return box, length, width, (float(cx), float(cy))


def mask_centroid_xy(mask_bool: np.ndarray):
    ys, xs = np.where(mask_bool)
    if len(xs) == 0:
        return None
    return (float(xs.mean()), float(ys.mean()))


def inside_mask(mask_bool: np.ndarray, xy):
    if xy is None:
        return False
    x, y = int(round(xy[0])), int(round(xy[1]))
    h, w = mask_bool.shape
    if x < 0 or x >= w or y < 0 or y >= h:
        return False
    return bool(mask_bool[y, x])


def mask_iou(a_bool: np.ndarray, b_bool: np.ndarray) -> float:
    inter = np.logical_and(a_bool, b_bool).sum()
    if inter == 0:
        return 0.0
    union = np.logical_or(a_bool, b_bool).sum()
    return float(inter) / float(union) if union > 0 else 0.0


def safe_mean(arr: np.ndarray) -> float:
    return float(np.nan) if len(arr) == 0 else float(np.mean(arr))


def anchor_xy(d: Dict[str, Any]) -> Tuple[int, int]:
    a = d["mbr_center"] if d["mbr_center"] is not None else d["centroid"]
    if a is None:
        return (10, 10)
    return (int(round(a[0])), int(round(a[1])))


# ------------------------- Excel writing helpers ------------------------- #

def autosize_worksheet(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_vertical_kv_sheet(wb, sheet_name: str, kv_dict: Dict[str, Any], title: str = "Image Summary"):
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center")

    ws["A1"] = title
    ws["A1"].font = title_font

    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    for c in ("A3", "B3"):
        ws[c].fill = header_fill
        ws[c].font = header_font
        ws[c].alignment = center

    r = 4
    for k, v in kv_dict.items():
        ws.cell(row=r, column=1, value=str(k))
        ws.cell(row=r, column=2, value=v)
        r += 1

    ws.freeze_panes = "A4"
    autosize_worksheet(ws)
    return ws


def write_df_sheet(wb, sheet_name: str, df: pd.DataFrame, freeze_header: bool = True):
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    if freeze_header:
        ws.freeze_panes = "A2"

    autosize_worksheet(ws)
    return ws


def save_excel(out_path: Union[str, Path], summary: Dict[str, Any], inst_df: pd.DataFrame):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_vertical_kv_sheet(wb, "ImageSummary", summary, title="Image Summary")
    write_df_sheet(wb, "Instances", inst_df)
    wb.save(str(out_path))


def save_excel_batch(
    out_path: Union[str, Path],
    agg_summary: Dict[str, Any],
    per_image_df: pd.DataFrame,
    inst_df: pd.DataFrame,
):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_vertical_kv_sheet(wb, "AggregatedImageSummary", agg_summary, title="Aggregated Image Summary")
    write_df_sheet(wb, "PerImageSummary", per_image_df)
    write_df_sheet(wb, "Instances", inst_df)

    wb.save(str(out_path))


def build_aggregated_summary(per_image_df: pd.DataFrame) -> Dict[str, Any]:
    numeric_cols = [
        "image_area_µm²",
        "stomata_count_yolo",
        "pore_count_yolo",
        "stomata_density_per_mm²",
        "pore_density_per_mm²",
        "stomata_length_mean_µm",
        "stomata_width_mean_µm",
        "stomata_area_mean_µm²",
        "pore_length_mean_µm",
        "pore_width_mean_µm",
        "pore_area_mean_µm²",
    ]

    existing = [c for c in numeric_cols if c in per_image_df.columns]
    df_num = per_image_df[existing].apply(pd.to_numeric, errors="coerce")

    agg: Dict[str, Any] = {}
    agg["image_count"] = int(len(per_image_df))

    stom = per_image_df.get("stomata_count_yolo", pd.Series([0] * len(per_image_df))).fillna(0)
    pore = per_image_df.get("pore_count_yolo", pd.Series([0] * len(per_image_df))).fillna(0)
    agg["images_with_detections"] = int(((stom + pore) > 0).sum())
    agg["total_stomata_count_yolo"] = int(stom.sum())
    agg["total_pore_count_yolo"] = int(pore.sum())

    for c in existing:
        s = df_num[c].dropna()
        if len(s) == 0:
            agg[f"{c}__mean"] = np.nan
            agg[f"{c}__median"] = np.nan
            agg[f"{c}__std"] = np.nan
            agg[f"{c}__min"] = np.nan
            agg[f"{c}__max"] = np.nan
        else:
            agg[f"{c}__mean"] = float(s.mean())
            agg[f"{c}__median"] = float(s.median())
            agg[f"{c}__std"] = float(s.std(ddof=1)) if len(s) > 1 else 0.0
            agg[f"{c}__min"] = float(s.min())
            agg[f"{c}__max"] = float(s.max())

    return agg


# ------------------------- Pipeline class ------------------------- #

class StomataPorePipeline:
    """
    Load YOLO + SAM once, then call run_image/run_images/run_folder_batch repeatedly.
    Django note: SamPredictor is stateful; avoid multi-threaded access to one instance.
    """

    def __init__(
        self,
        yolo_weights: Union[str, Path],
        sam_ckpt: Union[str, Path],
        sam_type: str,
        cfg: PipelineConfig,
        scale: ScaleConfig,
        class_color: Optional[Dict[int, Tuple[int, int, int]]] = None,
        text_color: Tuple[int, int, int] = (255, 255, 255),
    ):
        self.cfg = cfg
        self.scale = scale
        self.class_color = class_color or {0: (0, 255, 0), 1: (0, 0, 255)}  # BGR
        self.text_color = text_color

        # YOLO
        self.yolo = YOLO(str(yolo_weights))

        # SAM
        self.sam = sam_model_registry[sam_type](checkpoint=str(sam_ckpt))

        self._yolo_device, self._sam_device = normalize_device(cfg.device)
        self.sam.to(device=self._sam_device)
        self.predictor = SamPredictor(self.sam)

    def run_image(
        self,
        image_path: Union[str, Path],
        out: Optional[OutputConfig] = None,
        image_id: Optional[str] = None,
    ) -> ImageResult:
        t0 = time.time()
        image_path = Path(image_path)
        image_name = image_id or image_path.name

        bgr = load_bgr(image_path)
        rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        H, W = rgb.shape[:2]

        um_per_px = float(self.scale.um_per_px)
        img_w_um = W * um_per_px
        img_h_um = H * um_per_px
        img_area_um2 = img_w_um * img_h_um
        img_area_mm2 = img_area_um2 / 1e6  # 1 mm² = 1e6 µm²

        overlay = bgr.copy()

        result = self.yolo.predict(
            source=rgb,
            imgsz=self.cfg.imgsz,
            rect=self.cfg.rect,
            conf=self.cfg.conf,
            iou=self.cfg.iou,
            device=self._yolo_device,
            amp=self.cfg.amp,
            verbose=self.cfg.verbose,
        )[0]

        # No detections
        if result.obb is None or result.obb.xyxyxyxy is None or len(result.obb.xyxyxyxy) == 0:
            summary = {
                "image": image_name,
                "H_px": H,
                "W_px": W,
                "µm_per_px": um_per_px,
                "image_area_µm²": img_area_um2,
                "stomata_count_yolo": 0,
                "pore_count_yolo": 0,
                "stomata_density_per_mm²": 0.0,
                "pore_density_per_mm²": 0.0,
                "stomata_length_mean_µm": np.nan,
                "stomata_width_mean_µm": np.nan,
                "stomata_area_mean_µm²": np.nan,
                "pore_length_mean_µm": np.nan,
                "pore_width_mean_µm": np.nan,
                "pore_area_mean_µm²": np.nan,
            }

            inst_df = pd.DataFrame(columns=[
                "image", "uid", "instance", "class_id", "confidence",
                "length_px", "width_px", "pixel_count", "length_µm", "width_µm", "area_µm²"
            ])

            overlay_path, excel_path = None, None
            if out is not None:
                overlay_path, excel_path = self._save_outputs(out, image_name, overlay, summary, inst_df)

            return ImageResult(
                image_name=image_name,
                summary=summary,
                instances_df=inst_df,
                overlay_bgr=overlay,
                overlay_path=overlay_path,
                excel_path=excel_path,
                elapsed_sec=time.time() - t0,
            )

        # SAM for all detections
        self.predictor.set_image(rgb)

        instances: List[Dict[str, Any]] = []
        obb_xy = result.obb.xyxyxyxy.cpu().numpy()
        obb_cls = result.obb.cls.cpu().numpy().astype(int)
        obb_conf = result.obb.conf.cpu().numpy()

        raw_stomata_count = int((obb_cls == 0).sum())
        raw_pore_count = int((obb_cls == 1).sum())

        for i, (pts, cls_id, conf) in enumerate(zip(obb_xy, obb_cls, obb_conf)):
            x0, y0 = pts.min(axis=0)
            x1, y1 = pts.max(axis=0)
            box_prompt = np.array([x0, y0, x1, y1], dtype=np.float32)

            masks, _, _ = self.predictor.predict(box=box_prompt, multimask_output=False)
            sam_mask = masks[0].astype(bool)

            # "Fail-safe": SAM clipped by YOLO OBB polygon
            tight_mask = sam_mask & obb_polygon_mask(pts, H, W)

            area_px = int(tight_mask.sum())
            mbr_pts, length_px, width_px, mbr_center = mask_to_mbr(tight_mask)
            centroid = mask_centroid_xy(tight_mask)

            instances.append({
                "instance": i,
                "class_id": int(cls_id),
                "confidence": float(conf),
                "mask": tight_mask,
                "area_px": area_px,
                "length_px": float(length_px),
                "width_px": float(width_px),
                "mbr_pts": mbr_pts,
                "mbr_center": mbr_center,
                "centroid": centroid,
            })

        stomata = [d for d in instances if d["class_id"] == 0 and d["area_px"] > 0]
        pores_all = [d for d in instances if d["class_id"] == 1 and d["area_px"] > 0]

        # Assign UIDs to stomata
        for uid, s in enumerate(stomata, start=1):
            s["uid"] = uid

        # Strict biological rule:
        # - a pore must be contained in a stomata; otherwise drop the pore
        valid_pores: List[Dict[str, Any]] = []
        for p in pores_all:
            if not stomata:
                continue  # no stomata => no valid pores

            p_pt = p["centroid"] if p["centroid"] is not None else p["mbr_center"]
            if p_pt is None:
                continue

            containing = [s for s in stomata if inside_mask(s["mask"], p_pt)]

            if len(containing) == 1:
                p["uid"] = containing[0]["uid"]
                valid_pores.append(p)
            elif len(containing) > 1:
                # if overlap, choose max IoU among containing masks
                best_uid, best_iou = max(
                    ((s["uid"], mask_iou(p["mask"], s["mask"])) for s in containing),
                    key=lambda x: x[1],
                )
                if best_iou > 0.0:
                    p["uid"] = best_uid
                    valid_pores.append(p)
            else:
                # not contained in any stomata => invalid pore
                continue

        pores = valid_pores

        # Rebuild instances so "Instances" sheet only includes valid pores
        instances = [d for d in stomata] + [d for d in pores]

        stomata_count_final = len(stomata)
        pore_count_final = len(pores)

        # Instances table
        inst_rows = []
        for d in instances:
            if d["area_px"] <= 0:
                continue
            uid = int(d.get("uid", -1))
            length_um = d["length_px"] * um_per_px
            width_um = d["width_px"] * um_per_px
            area_um2 = d["area_px"] * (um_per_px ** 2)

            inst_rows.append({
                "image": image_name,
                "uid": uid,
                "instance": int(d["instance"]),
                "class_id": int(d["class_id"]),
                "confidence": float(d["confidence"]),
                "length_px": float(d["length_px"]),
                "width_px": float(d["width_px"]),
                "pixel_count": int(d["area_px"]),
                "length_µm": float(length_um),
                "width_µm": float(width_um),
                "area_µm²": float(area_um2),
            })

        inst_df = pd.DataFrame(inst_rows).sort_values(["uid", "class_id", "instance"], kind="stable")

        st0 = inst_df[inst_df["class_id"] == 0]
        st1 = inst_df[inst_df["class_id"] == 1]

        summary = {
            "image": image_name,
            "H_px": H,
            "W_px": W,
            "µm_per_px": um_per_px,
            "image_area_µm²": img_area_um2,

            # Final reported counts (after pore filtering)
            "stomata_count_yolo": stomata_count_final,
            "pore_count_yolo": pore_count_final,

            "stomata_density_per_mm²": (stomata_count_final / img_area_mm2) if img_area_mm2 > 0 else np.nan,
            "pore_density_per_mm²": (pore_count_final / img_area_mm2) if img_area_mm2 > 0 else np.nan,

            "stomata_length_mean_µm": safe_mean(st0["length_µm"].to_numpy()),
            "stomata_width_mean_µm": safe_mean(st0["width_µm"].to_numpy()),
            "stomata_area_mean_µm²": safe_mean(st0["area_µm²"].to_numpy()),
            "pore_length_mean_µm": safe_mean(st1["length_µm"].to_numpy()),
            "pore_width_mean_µm": safe_mean(st1["width_µm"].to_numpy()),
            "pore_area_mean_µm²": safe_mean(st1["area_µm²"].to_numpy()),

            # optional debug
            "raw_stomata_count_yolo": raw_stomata_count,
            "raw_pore_count_yolo": raw_pore_count,
            "dropped_pore_count": int(max(0, raw_pore_count - pore_count_final)),
        }

        # Overlay drawing (only valid instances)
        uid_to: Dict[int, List[Dict[str, Any]]] = {}
        for d in instances:
            if d["area_px"] <= 0:
                continue
            uid = int(d.get("uid", -1))
            uid_to.setdefault(uid, []).append(d)

        for d in instances:
            if d["area_px"] <= 0:
                continue
            cls_id = d["class_id"]
            col = self.class_color.get(cls_id, (255, 255, 255))

            mask_u8 = (d["mask"].astype(np.uint8) * 255)
            contours, _ = cv2.findContours(mask_u8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
            if contours:
                cv2.drawContours(overlay, contours, -1, color=col, thickness=1, lineType=cv2.LINE_8)

            if d["mbr_pts"] is not None:
                box_i = np.round(d["mbr_pts"]).astype(np.int32).reshape(-1, 1, 2)
                cv2.polylines(overlay, [box_i], isClosed=True, color=col, thickness=1, lineType=cv2.LINE_8)

        for uid, items in uid_to.items():
            if uid == -1:
                continue
            pores_here = [d for d in items if d["class_id"] == 1]
            stom_here = [d for d in items if d["class_id"] == 0]
            ref = pores_here[0] if pores_here else (stom_here[0] if stom_here else items[0])
            ax, ay = anchor_xy(ref)
            cv2.putText(
                overlay, f"ID:{uid}",
                (ax + 3, max(15, ay - 3)),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                self.text_color, 1, cv2.LINE_AA
            )

        overlay_path, excel_path = None, None
        if out is not None:
            overlay_path, excel_path = self._save_outputs(out, image_name, overlay, summary, inst_df)

        return ImageResult(
            image_name=image_name,
            summary=summary,
            instances_df=inst_df,
            overlay_bgr=overlay,
            overlay_path=overlay_path,
            excel_path=excel_path,
            elapsed_sec=time.time() - t0,
        )

    def run_images(
        self,
        image_paths: List[Union[str, Path]],
        out: OutputConfig,
        batch_excel_name: str = "batch_results.xlsx",
        save_per_image_excel: bool = False,
    ) -> Dict[str, Any]:
        """
        Batch runner:
        - Saves overlays per image (based on out.save_overlay)
        - Produces ONE Excel: AggregatedImageSummary + PerImageSummary + Instances
        - Optionally saves per-image Excel if save_per_image_excel=True
        """
        out_dir = Path(out.out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        per_image_rows: List[Dict[str, Any]] = []
        all_instances: List[pd.DataFrame] = []
        overlay_paths: Dict[str, str] = {}

        for p in image_paths:
            p = Path(p)

            per_img_out = OutputConfig(
                out_dir=out.out_dir,
                save_overlay=out.save_overlay,
                save_excel=save_per_image_excel,
                overlay_suffix=out.overlay_suffix,
                excel_suffix=out.excel_suffix,
            )

            res = self.run_image(p, out=per_img_out)

            per_image_rows.append(res.summary)
            all_instances.append(res.instances_df)

            if res.overlay_path:
                overlay_paths[res.image_name] = res.overlay_path

        per_image_df = pd.DataFrame(per_image_rows)

        if len(all_instances) > 0:
            inst_df = pd.concat(all_instances, ignore_index=True)
        else:
            inst_df = pd.DataFrame(columns=[
                "image", "uid", "instance", "class_id", "confidence",
                "length_px", "width_px", "pixel_count", "length_µm", "width_µm", "area_µm²"
            ])

        agg_summary = build_aggregated_summary(per_image_df)

        batch_excel_path = str(out_dir / batch_excel_name)
        save_excel_batch(batch_excel_path, agg_summary, per_image_df, inst_df)

        return {
            "batch_excel_path": batch_excel_path,
            "overlay_paths": overlay_paths,
            "aggregated_summary": agg_summary,
            "per_image_df": per_image_df,
            "instances_df": inst_df,
        }

    def run_folder_batch(
        self,
        folder: Union[str, Path],
        out: OutputConfig,
        glob_pattern: str = "*.tif",
        batch_excel_name: str = "batch_results.xlsx",
        save_per_image_excel: bool = False,
    ) -> Dict[str, Any]:
        folder = Path(folder)
        image_paths = sorted(folder.glob(glob_pattern))
        return self.run_images(
            image_paths=[str(p) for p in image_paths],
            out=out,
            batch_excel_name=batch_excel_name,
            save_per_image_excel=save_per_image_excel,
        )

    def _save_outputs(
        self,
        out: OutputConfig,
        image_name: str,
        overlay_bgr: np.ndarray,
        summary: Dict[str, Any],
        inst_df: pd.DataFrame,
    ) -> Tuple[Optional[str], Optional[str]]:
        out_dir = Path(out.out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        base = Path(image_name).stem

        overlay_path = None
        if out.save_overlay:
            overlay_path = str(out_dir / f"{base}{out.overlay_suffix}")
            cv2.imwrite(overlay_path, overlay_bgr)

        excel_path = None
        if out.save_excel:
            excel_path = str(out_dir / f"{base}{out.excel_suffix}")
            save_excel(excel_path, summary, inst_df)

        return overlay_path, excel_path


# ------------------------- Convenience for Django/JSON ------------------------- #

def df_to_records(df: pd.DataFrame) -> List[Dict[str, Any]]:
    return df.replace({np.nan: None}).to_dict(orient="records")
